import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Colour palette ──────────────────────────────────────────────────────────
DARK_NAVY   = "0D1B2A"   # header backgrounds
BLUE_HEADER = "1A3A5C"   # section headers
LIGHT_BLUE  = "D6E4F0"   # alternate row
WHITE       = "FFFFFF"
ORANGE      = "F5A623"   # GAP
GREEN       = "27AE60"   # Available / Config / Derivable
AMBER       = "E67E22"   # Mapping needed / May be null / Needs validation
RED_LIGHT   = "E74C3C"   # HIGH priority
AMBER_MED   = "F39C12"   # MEDIUM priority
GREEN_LOW   = "27AE60"   # LOW priority

def hex_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_font(bold=False, color=WHITE, size=10):
    return Font(bold=bold, color=color, name="Calibri", size=size)

thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def header_row(ws, row, values, bg=DARK_NAVY, fg=WHITE, bold=True, sizes=None):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = hex_fill(bg)
        c.font = Font(bold=bold, color=fg, name="Calibri",
                      size=(sizes[col-1] if sizes else 10))
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        c.border = border

def data_row(ws, row, values, bg=WHITE, fg="000000", bold=False, wrap=True):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = hex_fill(bg)
        c.font = Font(bold=bold, color=fg, name="Calibri", size=10)
        c.alignment = Alignment(wrap_text=wrap, vertical="top")
        c.border = border

def alt_bg(row_idx):
    return LIGHT_BLUE if row_idx % 2 == 0 else WHITE

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def status_cell(ws, row, col, status_text):
    """Colour-code the status cell based on keyword."""
    c = ws.cell(row=row, column=col, value=status_text)
    s = status_text.lower()
    if "gap" in s:
        bg, fg = ORANGE, WHITE
    elif any(k in s for k in ["available", "config", "derivable", "hardcoded", "not required"]):
        bg, fg = GREEN, WHITE
    elif any(k in s for k in ["mapping needed", "may be", "needs validation", "incomplete"]):
        bg, fg = AMBER, WHITE
    else:
        bg, fg = WHITE, "000000"
    c.fill = hex_fill(bg)
    c.font = Font(bold=True, color=fg, name="Calibri", size=10)
    c.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")
    c.border = border

def priority_cell(ws, row, col, priority_text):
    c = ws.cell(row=row, column=col, value=priority_text)
    p = priority_text.upper()
    if "HIGH" in p:
        bg, fg = RED_LIGHT, WHITE
    elif "MEDIUM" in p:
        bg, fg = AMBER_MED, WHITE
    else:
        bg, fg = GREEN_LOW, WHITE
    c.fill = hex_fill(bg)
    c.font = Font(bold=True, color=fg, name="Calibri", size=10)
    c.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")
    c.border = border

def section_title(ws, row, title, span, bg=BLUE_HEADER):
    c = ws.cell(row=row, column=1, value=title)
    c.fill = hex_fill(bg)
    c.font = Font(bold=True, color=WHITE, name="Calibri", size=11)
    c.alignment = Alignment(wrap_text=True, vertical="center")
    c.border = border
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 22


# ════════════════════════════════════════════════════════════════════════════
# SHEET 1 — OVERVIEW
# ════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "1. Overview"
ws1.sheet_view.showGridLines = False
ws1.row_dimensions[1].height = 40

# Title banner
ws1.merge_cells("A1:D1")
c = ws1["A1"]
c.value = "TenXEngage  ×  XTRM API — Integration Mapping Overview"
c.fill = hex_fill(DARK_NAVY)
c.font = Font(bold=True, color=WHITE, name="Calibri", size=14)
c.alignment = Alignment(horizontal="center", vertical="center")

header_row(ws1, 2, ["Property", "Value", "Notes", "Classification"], bg=BLUE_HEADER)

doc_info = [
    ("Vendor",           "XTRM (Cash Payouts)", "External payout provider for CASH redemptions", ""),
    ("Version",          "1.0 — April 2026", "", ""),
    ("Classification",   "HIGHLY CONFIDENTIAL", "Internal use only", ""),
    ("Auth Method",      "OAuth 2.0 Bearer Token", "POST /oAuth/token — cache token with expiry", ""),
    ("Sandbox Base URL", "https://xapisandbox.xtrm.com/API/V4/", "Use for all dev/test payout calls", ""),
    ("Production Base URL","https://xapi.xtrm.com/API/V4/", "Switch at go-live", ""),
    ("Webhook Endpoint", "POST /api/v1/redemption/webhook/xtrm", "Inbound from XTRM; HMAC-SHA256 verified", ""),
    ("Token Expiry",     "3600 seconds (1 hour)", "Cache access_token; use refresh_token before expiry", ""),
    ("", "", "", ""),
]

for i, row_data in enumerate(doc_info, 3):
    data_row(ws1, i, row_data, bg=alt_bg(i))

ws1.row_dimensions[10].height = 8

# Summary counts
section_title(ws1, 12, "Document Summary", 4)
header_row(ws1, 13, ["Section", "Count", "Description", "Status"], bg=BLUE_HEADER)
summary = [
    ("Part 1 — Data Model Entities",     "6",  "RewardWallet, LedgerEntry, RedemptionCatalogItem, ClientRedemptionConfig, RedemptionTransaction, RedemptionReturn", ""),
    ("Part 2 — Service Classes",         "4",  "WalletService, XtrmAdapter, XtrmWebhookService, RedemptionCatalogService / ReturnsService", ""),
    ("Part 3 — XTRM API Calls",         "10", "Auth + Beneficiary + Register + Wallet + Fund APIs", ""),
    ("Part 4 — Data Gaps",              "12", "Fields missing from current tenXengage data model", "Action required"),
    ("Part 5 — Open Questions",         "12", "Ranging HIGH → LOW priority", "7 HIGH, 4 MEDIUM, 1 LOW"),
    ("Part 6 — REST Controllers",        "6",  "XTRM-relevant API endpoints", ""),
]
for i, row_data in enumerate(summary, 14):
    data_row(ws1, i, row_data, bg=alt_bg(i))

set_col_widths(ws1, [28, 40, 60, 22])


# ════════════════════════════════════════════════════════════════════════════
# SHEET 2 — DATA MODEL
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("2. Data Model")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:F1")
c = ws2["A1"]
c.value = "PART 1 — Data Model Entities"
c.fill = hex_fill(DARK_NAVY)
c.font = Font(bold=True, color=WHITE, name="Calibri", size=13)
c.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 36

row = 2
entities = [
    {
        "name": "RewardWallet",
        "scope": "Tenant-scoped (TenantAware + clientId + @Filter). Internal state only — no XTRM API reads from this entity. availableBalance and reservedBalance updated by XTRM webhook events.",
        "xtrm": "No direct XTRM calls — driven by webhook events",
        "fields": [
            ("ownerType",         "Enum",       "INDIVIDUAL | COMPANY",                      "Required for payout routing"),
            ("userId",            "UUID",        "Nullable FK → users",                       "Individual owner"),
            ("partnerCompanyId",  "UUID",        "Nullable FK → partner_companies",           "Company owner"),
            ("currencyType",      "Enum",        "CASH | POINTS | CREDITS | TICKETS",         "CASH routes to XTRM"),
            ("availableBalance",  "BigDecimal",  "NUMERIC(19,4) — updated by XTRM webhook",   "Decremented on SUCCEEDED"),
            ("reservedBalance",   "BigDecimal",  "NUMERIC(19,4) — updated by XTRM webhook",   "Incremented on reserve"),
            ("totalEarned",       "BigDecimal",  "NUMERIC(19,4)",                             "Cumulative"),
            ("[inherited]",       "—",           "id UUID, clientId UUID, createdAt, updatedAt", ""),
        ]
    },
    {
        "name": "LedgerEntry",
        "scope": "Tenant-scoped. Immutable after creation. Written by WalletService in response to XTRM webhook events.",
        "xtrm": "Written BY webhook events (SUCCEEDED → DEBIT; FAILED/CANCELLED → RELEASE)",
        "fields": [
            ("walletId",              "UUID",       "FK → reward_wallet",                        ""),
            ("entryType",             "Enum",       "CREDIT | DEBIT | RESERVE | RELEASE | RETURN_CREDIT | EXPIRY_DEBIT", ""),
            ("sourceType",            "Enum",       "REWARD_EARNED | REDEMPTION_RESERVE | REDEMPTION_DEBIT | REDEMPTION_RELEASE | RETURN_CREDIT | ADJUSTMENT | CLAWBACK", ""),
            ("currencyType",          "CurrencyType","",                                         ""),
            ("amount",                "BigDecimal", "Always positive, CHECK > 0",                ""),
            ("relatedTransactionId",  "UUID",       "Nullable — links to RedemptionTransaction", ""),
            ("[inherited]",           "—",          "id UUID, clientId UUID, createdAt, updatedAt", ""),
        ]
    },
    {
        "name": "RedemptionCatalogItem",
        "scope": "Platform-level — NOT tenant-scoped (no clientId, no @Filter). Items with category=CASH route to XTRM. providerItemId NOT used for CASH items.",
        "xtrm": "CASH items manually seeded — XTRM has no catalog API",
        "fields": [
            ("name",                  "String",  "",                                             ""),
            ("category",              "Enum",    "CASH | NON_CASH — CASH routes to XTRM",       "Key routing field"),
            ("subCategory",           "String",  "",                                             ""),
            ("currencyType",          "CurrencyType", "",                                        ""),
            ("minimumAmount",         "BigDecimal", "",                                          ""),
            ("defaultProcessingMode", "Enum",    "INSTANT | BATCH | APPROVAL_REQUIRED",          ""),
            ("providerItemId",        "String",  "NOT USED for CASH/XTRM items (Xoxoday only)",  "Leave null for CASH"),
            ("geographicScope",       "String[]","TEXT[]",                                       ""),
            ("isReturnable",          "Boolean", "Always false for CASH items",                  ""),
            ("isActive",              "Boolean", "",                                              ""),
            ("[inherited]",           "—",       "id UUID, createdAt, updatedAt (no clientId)",  ""),
        ]
    },
    {
        "name": "ClientRedemptionConfig",
        "scope": "Tenant-scoped. processingModeOverride determines XTRM call timing (INSTANT/BATCH/APPROVAL_REQUIRED). minimumTransactionAmount gates minimum sent to XTRM.",
        "xtrm": "Controls whether/when XTRM is called",
        "fields": [
            ("catalogItemId",            "UUID",       "FK → redemption_catalog_item",              ""),
            ("isEnabled",                "Boolean",    "Must be true for XTRM to be called",         "Hard gate"),
            ("processingModeOverride",   "Enum",       "Nullable — overrides catalog default",       ""),
            ("minimumWalletBalance",     "BigDecimal", "Wallet must exceed this before redemption",  ""),
            ("minimumTransactionAmount", "BigDecimal", "XTRM payout must be at least this amount",  ""),
            ("returnWindowDays",         "Integer",    "N/A for CASH (returns blocked)",              ""),
            ("[inherited]",              "—",          "id UUID, clientId UUID, createdAt, updatedAt",""),
        ]
    },
    {
        "name": "RedemptionTransaction",
        "scope": "Tenant-scoped. Central XTRM entity. id → XTRM IssuerTransactionId. XTRM TransactionId stored back in providerTransactionId. Status lifecycle driven by XTRM webhooks.",
        "xtrm": "id sent as IssuerTransactionId; XTRM TransactionId stored in providerTransactionId",
        "fields": [
            ("walletId",             "UUID",       "FK → reward_wallet",                         ""),
            ("catalogItemId",        "UUID",       "FK → redemption_catalog_item",               ""),
            ("requestedBy",          "UUID",       "FK → users — used to resolve XTRM recipient",""),
            ("amount",               "BigDecimal", "Sent as PaymentAmount to XTRM",              ""),
            ("currencyType",         "CurrencyType","Mapped to ISO 4217 for XTRM",               ""),
            ("status",               "Enum",       "TransactionStatus platform enum",            "Driven by XTRM webhook"),
            ("processingMode",       "Enum",       "INSTANT | BATCH | APPROVAL_REQUIRED",        ""),
            ("providerName",         "Enum",       "XTRM | XOXODAY",                             ""),
            ("providerTransactionId","String",     "Nullable — XTRM TransactionId stored here",  "Set on webhook"),
            ("approvedBy",           "UUID",       "Nullable",                                   ""),
            ("approvedAt",           "Instant",    "Nullable",                                   ""),
            ("completedAt",          "Instant",    "Nullable — set on XTRM SUCCEEDED webhook",   ""),
            ("failureReason",        "String",     "Nullable — set on XTRM FAILED webhook",      ""),
            ("returnId",             "UUID",       "Nullable — N/A for CASH/XTRM",               ""),
            ("[inherited]",          "—",          "id UUID, clientId UUID, createdAt, updatedAt",""),
        ]
    },
    {
        "name": "RedemptionReturn",
        "scope": "NO XTRM involvement. Cash (XTRM) redemptions are explicitly non-returnable. ReturnsService.submitReturnRequest() throws BusinessRuleException if category=CASH.",
        "xtrm": "NO XTRM API calls — returns blocked for CASH",
        "fields": [
            ("redemptionTransactionId","UUID",    "Unique FK → redemption_transaction",           ""),
            ("requestedBy",           "UUID",    "FK → users",                                   ""),
            ("status",                "Enum",    "PENDING_APPROVAL | APPROVED | RETURN_CONFIRMED | RETURN_REJECTED | CANCELLED", ""),
            ("returnReason",          "String",  "Nullable",                                     ""),
            ("reviewedBy",            "UUID",    "Nullable",                                     ""),
            ("reviewedAt",            "Instant", "Nullable",                                     ""),
            ("providerReturnId",      "String",  "Xoxoday only — never XTRM",                    ""),
            ("providerConfirmedAt",   "Instant", "Nullable",                                     ""),
            ("walletCreditedAt",      "Instant", "Nullable",                                     ""),
            ("rejectionReason",       "String",  "Nullable",                                     ""),
            ("[inherited]",           "—",       "id UUID, clientId UUID, createdAt, updatedAt", ""),
        ]
    },
]

for entity in entities:
    # Entity section header
    section_title(ws2, row, f"Entity: {entity['name']}", 6)
    row += 1
    # Scope note
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    c = ws2.cell(row=row, column=1, value=f"Scope/Notes: {entity['scope']}")
    c.fill = hex_fill("EBF5FB")
    c.font = Font(italic=True, color="1A3A5C", name="Calibri", size=9)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    c.border = border
    ws2.row_dimensions[row].height = 30
    row += 1
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    c = ws2.cell(row=row, column=1, value=f"XTRM Involvement: {entity['xtrm']}")
    c.fill = hex_fill("FDFEFE")
    c.font = Font(bold=True, italic=True, color="117A65", name="Calibri", size=9)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    c.border = border
    row += 1
    header_row(ws2, row, ["Field Name", "Type", "Value / Enum", "Notes", "", ""], bg="2E4057")
    ws2.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
    row += 1
    for fi, (fname, ftype, fval, fnotes) in enumerate(entity["fields"]):
        bg = alt_bg(fi)
        ws2.cell(row=row, column=1, value=fname).fill = hex_fill(bg)
        ws2.cell(row=row, column=2, value=ftype).fill = hex_fill(bg)
        ws2.cell(row=row, column=3, value=fval).fill = hex_fill(bg)
        ws2.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
        ws2.cell(row=row, column=4, value=fnotes).fill = hex_fill(bg)
        for col in range(1, 7):
            c = ws2.cell(row=row, column=col)
            c.font = Font(name="Calibri", size=10, color="000000")
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border = border
        ws2.cell(row=row, column=1).font = Font(name="Courier New", size=9, color="1A5276")
        row += 1
    ws2.row_dimensions[row].height = 8
    row += 1

set_col_widths(ws2, [28, 14, 50, 20, 10, 10])


# ════════════════════════════════════════════════════════════════════════════
# SHEET 3 — XTRM API INVENTORY
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("3. API Inventory")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:E1")
c = ws3["A1"]
c.value = "PART 3 — Full XTRM API Call Inventory"
c.fill = hex_fill(DARK_NAVY)
c.font = Font(bold=True, color=WHITE, name="Calibri", size=13)
c.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 36

header_row(ws3, 2, ["#", "XTRM Endpoint", "Method", "Called From", "Trigger / Notes"], bg=BLUE_HEADER)

apis = [
    ("1",  "/oAuth/token",                          "POST", "XtrmAdapter",                          "App startup + token expiry. Cache access_token (3600s). Use refresh_token before expiry."),
    ("2",  "/API/V4/Beneficiary/CheckUserExist",    "POST", "XtrmAdapter.submitPayout()",            "Every payout — check cache first. If UserExists=true, skip Step 2b."),
    ("3",  "/API/V4/Register/CreateUser",           "POST", "XtrmAdapter.submitPayout()",            "Only when user NOT found in Step 2. Store returned UserId as User.xtrmUserId."),
    ("4",  "/API/V4/Wallet/CreateUserWallet",       "POST", "WalletService.getWallet()",             "Lazy init of individual user XTRM wallet. Called once per user. Store WalletID as User.xtrmWalletId."),
    ("5",  "/API/V4/Wallet/CreateCompanyWallet",    "POST", "WalletService.getCompanyWallet()",      "Lazy init of company XTRM wallet. Store WalletID as PartnerCompany.xtrmWalletId."),
    ("6",  "/API/V4/Fund/TransferFund",             "POST", "XtrmAdapter.submitPayout()",            "INDIVIDUAL cash payout. ownerType=INDIVIDUAL. Async — status via webhook."),
    ("7",  "/API/V4/Fund/TransferFundtoCompany",    "POST", "XtrmAdapter.submitPayout()",            "COMPANY cash payout. ownerType=COMPANY. Requires BeneficiaryAccountNumber + BeneficiaryWalletID."),
    ("8",  "/API/V4/Wallet/GetCompanyWallets",      "POST", "Config init / admin tool",              "One-time setup: fetch source WalletID after XTRM account is funded. Store in config."),
    ("9",  "/API/V4/Programs/GetPrograms",          "POST", "Config init / admin tool",              "One-time setup: fetch ProgramId for Redemption Store program. Store in config."),
    ("10", "/API/V4/Payment/GetPaymentMethods",     "POST", "Config init / admin tool",              "One-time setup: fetch PaymentMethodId. Use XTRM Choice (XTR94503) for recipient flexibility."),
    ("—",  "(Inbound) POST /api/v1/redemption/webhook/xtrm", "—", "XtrmWebhookService",             "XTRM posts payout status (SUCCEEDED/FAILED/CANCELLED). HMAC-SHA256 verified. Always return HTTP 200."),
]

for i, row_data in enumerate(apis, 3):
    bg = alt_bg(i)
    for col, val in enumerate(row_data, 1):
        c = ws3.cell(row=i, column=col, value=val)
        c.fill = hex_fill(bg)
        c.font = Font(name="Calibri", size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.border = border
    # Highlight endpoint col
    ws3.cell(row=i, column=2).font = Font(name="Courier New", size=9, color="1A5276")

set_col_widths(ws3, [5, 48, 10, 38, 60])
for i in range(3, 14):
    ws3.row_dimensions[i].height = 42


# ════════════════════════════════════════════════════════════════════════════
# SHEET 4 — FIELD MAPPINGS (all API calls)
# ════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("4. Field Mappings")
ws4.sheet_view.showGridLines = False

ws4.merge_cells("A1:F1")
c = ws4["A1"]
c.value = "PART 2 — Service + XTRM API Field Mappings (all API calls)"
c.fill = hex_fill(DARK_NAVY)
c.font = Font(bold=True, color=WHITE, name="Calibri", size=13)
c.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 36

row = 2

field_sections = [
    {
        "api": "POST /oAuth/token  (XtrmAdapter — Step 1: Authentication)",
        "cols": ["XTRM Field", "tenXengage Source", "Notes"],
        "rows": [
            ("client_id",     "@Value(xtrm.client-id)",     "Never log — keep in secrets manager"),
            ("client_secret", "@Value(xtrm.client-secret)", "Never log"),
            ("grant_type",    'Hardcoded "password"',        "Static constant"),
        ],
        "status": None
    },
    {
        "api": "POST /API/V4/Beneficiary/CheckUserExist  (XtrmAdapter — Step 2: Recipient Lookup)",
        "cols": ["XTRM Field", "tenXengage Source", "Status", "Notes"],
        "rows": [
            ("IssuerAccountNumber", "@Value(xtrm.issuer-account-number)", "Config",     "Static"),
            ("Email",               "User.email (resolved from tx.requestedBy UUID)", "Available", "Unique lookup key on XTRM side"),
        ],
        "status": 2
    },
    {
        "api": "POST /API/V4/Register/CreateUser  (XtrmAdapter — Step 2b: Recipient Registration — only if not found)",
        "cols": ["XTRM Field", "tenXengage Source", "Status", "How to Obtain if Missing"],
        "rows": [
            ("IssuerAccountNumber",       "@Value(xtrm.issuer-account-number)", "Available",           "Static config"),
            ("LegalFirstName",            "User.firstName",                      "Available",           "Direct mapping"),
            ("LegalLastName",             "User.lastName",                       "Available",           "Direct mapping"),
            ("EmailAddress",              "User.email",                          "Available",           "Direct mapping"),
            ("MobilePhone",               "User.phoneNumber",                    "May be null",         "Optional — send if present, skip if null"),
            ("TaxId",                     "User.taxId",                          "GAP",                 "Add encrypted taxId to UserProfile. Collect at onboarding/first CASH redemption via KYC. Consider XTRM native tax form flow."),
            ("DateOfBirth (Day/Mon/Year)","User.dateOfBirth",                    "GAP",                 "Add dateOfBirth (LocalDate) to UserProfile. Collect at onboarding."),
            ("Address.AddressLine1",      "User.address.line1",                  "May be incomplete",   "Validate completeness at first CASH redemption. Prompt user if missing."),
            ("Address.CountryISO2",       "User.address.countryCode",            "Needs validation",    "Ensure ISO 3166-1 alpha-2 stored on user"),
        ],
        "status": 2
    },
    {
        "api": "POST /API/V4/Wallet/CreateUserWallet  (WalletService — lazy init of individual wallet)",
        "cols": ["XTRM Field", "tenXengage Source", "Status", "Notes"],
        "rows": [
            ("IssuerAccountNumber", "@Value(xtrm.issuer-account-number)", "Config",          "Provided by XTRM at onboarding"),
            ("UserID",              "User.xtrmUserId",                     "GAP",             "Add xtrmUserId field to users table. Populated by CheckUserExist/CreateUser flow."),
            ("WalletName",          '"tenXengage-CASH-" + userId',         "Derivable",       "Constructed — must be unique per user"),
            ("WalletCurrency",      'CurrencyType.CASH → "USD"',           "Mapping needed",  "Add static map or cashIsoCurrencyCode to ClientRedemptionConfig"),
        ],
        "status": 2
    },
    {
        "api": "POST /API/V4/Wallet/CreateCompanyWallet  (WalletService — lazy init of company wallet)",
        "cols": ["XTRM Field", "tenXengage Source", "Status", "Notes"],
        "rows": [
            ("IssuerAccountNumber",    "@Value(xtrm.issuer-account-number)", "Config",          "Static config"),
            ("WalletName",             '"tenXengage-COMPANY-CASH-" + partnerCompanyId', "Derivable", "Must be unique"),
            ("WalletCurrency",         'CurrencyType.CASH → "USD"',          "Mapping needed",  "Same mapping as user wallet"),
            ("WalletType",             '"Standard"',                          "Hardcoded",       'Confirm with XTRM if "Accrual" needed'),
            ("AllowAccessAccountNumber","null",                               "Not required",    "Leave null for payout-only flow"),
        ],
        "status": 2
    },
    {
        "api": "POST /API/V4/Fund/TransferFund  (XtrmAdapter — Step 3: Individual Cash Payout)",
        "cols": ["XTRM Field", "tenXengage Source", "Status", "Notes"],
        "rows": [
            ("IssuerAccountNumber",    "@Value(xtrm.issuer-account-number)",                 "Config",          "Provided by XTRM at onboarding"),
            ("PaymentType",            'wallet.ownerType == INDIVIDUAL ? "Personal" : "Company"', "Derivable",   "Simple enum mapping"),
            ("PaymentMethodId",        "@Value(xtrm.payment-method-id)",                     "GAP",             "Fetch via GetUserPaymentMethods. Use XTRM Choice (XTR94503) for recipient flexibility."),
            ("ProgramId",              "@Value(xtrm.program-id)",                            "GAP",             "Create Redemption program in XTRM portal. Fetch via GetPrograms. Store in config."),
            ("WalletID",               "@Value(xtrm.company-wallet-id)",                     "GAP",             "Source wallet funding all payouts. Fetch via GetCompanyWallets after account is funded."),
            ("PaymentDescription",     '"Redemption: " + catalogItem.name',                  "Derivable",       "Dynamic"),
            ("PaymentCurrency",        'CurrencyType.CASH → "USD"',                          "Mapping needed",  "Add cashIsoCurrencyCode to ClientRedemptionConfig (default USD)"),
            ("IssuerTransactionId",    "tx.id.toString()",                                   "Available",       "CRITICAL — must be unique. Used for webhook reconciliation."),
            ("PaymentAmount",          "tx.amount.setScale(2,HALF_UP).toPlainString()",       "Available",       "BigDecimal → 2dp String"),
            ("RecipientUserId",        "User.xtrmUserId",                                    "GAP",             "Requires xtrmUserId field on User entity (from Step 2/2b)"),
            ("UserPrepaidVisaEmailID", "User.email",                                         "Available",       "Enables Xtrm Choice — recipient picks bank/Visa/PayPal"),
            ("Comment",                '"tenXengage clientId=" + tx.clientId',               "Derivable",       "Audit trail"),
        ],
        "status": 2
    },
    {
        "api": "POST /API/V4/Fund/TransferFundtoCompany  (XtrmAdapter — Step 3b: Company Cash Payout)",
        "cols": ["XTRM Field", "tenXengage Source", "Status", "Notes"],
        "rows": [
            ("IssuerAccountNumber",      "@Value(xtrm.issuer-account-number)", "Config",      ""),
            ("PaymentType",              '"Company"',                           "Hardcoded",   ""),
            ("PaymentMethodId",          "@Value(xtrm.payment-method-id)",     "GAP",         "Same gap as individual payout"),
            ("ProgramId",                "@Value(xtrm.program-id)",            "GAP",         ""),
            ("WalletID",                 "@Value(xtrm.company-wallet-id)",     "GAP",         ""),
            ("IssuerTransactionId",      "tx.id.toString()",                   "Available",   ""),
            ("BeneficiaryAccountNumber", "PartnerCompany.xtrmAccountNumber",   "GAP",         "Add xtrmAccountNumber to partner_companies table. Obtain from XTRM at company registration."),
            ("BeneficiaryWalletID",      "PartnerCompany.xtrmWalletId",        "GAP",         "Add xtrmWalletId to partner_companies. Returned from CreateCompanyWallet."),
        ],
        "status": 2
    },
]

for section in field_sections:
    section_title(ws4, row, section["api"], 6, bg=BLUE_HEADER)
    row += 1
    header_row(ws4, row, section["cols"] + [""]*(6-len(section["cols"])), bg="2E4057")
    if len(section["cols"]) < 6:
        ws4.merge_cells(start_row=row, start_column=len(section["cols"]), end_row=row, end_column=6)
    row += 1

    for fi, frow in enumerate(section["rows"]):
        bg = alt_bg(fi)
        for col_idx, val in enumerate(frow, 1):
            if section["status"] is not None and col_idx == section["status"] + 1:
                status_cell(ws4, row, col_idx, val)
            else:
                c = ws4.cell(row=row, column=col_idx, value=val)
                c.fill = hex_fill(bg)
                c.font = Font(name="Calibri", size=10)
                c.alignment = Alignment(wrap_text=True, vertical="top")
                c.border = border
        # Fill remaining cols
        for col_idx in range(len(frow)+1, 7):
            c = ws4.cell(row=row, column=col_idx)
            c.fill = hex_fill(bg)
            c.border = border
        ws4.cell(row=row, column=1).font = Font(name="Courier New", size=9, color="1A5276")
        row += 1
    ws4.row_dimensions[row].height = 6
    row += 1

set_col_widths(ws4, [32, 48, 20, 62, 5, 5])


# ════════════════════════════════════════════════════════════════════════════
# SHEET 5 — WEBHOOK STATUS MAPPING
# ════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("5. Webhook & Flow")
ws5.sheet_view.showGridLines = False

ws5.merge_cells("A1:E1")
c = ws5["A1"]
c.value = "XTRM Webhook — Status Mapping & Processing Steps"
c.fill = hex_fill(DARK_NAVY)
c.font = Font(bold=True, color=WHITE, name="Calibri", size=13)
c.alignment = Alignment(horizontal="center", vertical="center")
ws5.row_dimensions[1].height = 36

section_title(ws5, 2, "Webhook Endpoint: POST /api/v1/redemption/webhook/xtrm  |  Auth: permitAll (HMAC-SHA256 verified in handler)", 5, bg=BLUE_HEADER)

header_row(ws5, 3, ["XTRM Status", "tenXengage tx.status", "WalletService Call", "Additional Actions", "Domain Event"], bg="2E4057")
wh_rows = [
    ("SUCCEEDED", "COMPLETED", "walletService.debit(walletId, amount, currencyType, txId)", "Set completedAt = now()", "Publish COMPLETED domain event → notification service"),
    ("FAILED",    "FAILED",    "walletService.release(walletId, amount, currencyType, txId)","Set failureReason on tx", "Publish FAILED domain event → notification service"),
    ("CANCELLED", "CANCELLED", "walletService.release(walletId, amount, currencyType, txId)","—",                      "Publish CANCELLED domain event"),
]
status_colors = {"SUCCEEDED": ("27AE60", WHITE), "FAILED": (RED_LIGHT, WHITE), "CANCELLED": (AMBER_MED, WHITE)}
for i, (xstatus, txstatus, wallet_call, extra, event) in enumerate(wh_rows, 4):
    bg_c, fg_c = status_colors[xstatus]
    c = ws5.cell(row=i, column=1, value=xstatus)
    c.fill = hex_fill(bg_c); c.font = Font(bold=True, color=fg_c, name="Calibri", size=10)
    c.alignment = Alignment(horizontal="center", vertical="top"); c.border = border
    for col, val in enumerate([txstatus, wallet_call, extra, event], 2):
        cell = ws5.cell(row=i, column=col, value=val)
        cell.fill = hex_fill(alt_bg(i))
        cell.font = Font(name="Calibri", size=10)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = border
    ws5.row_dimensions[i].height = 35

ws5.row_dimensions[7].height = 10

section_title(ws5, 8, "Webhook Processing — Strict Order (8 Steps)", 5, bg=BLUE_HEADER)
header_row(ws5, 9, ["Step", "Action", "Detail", "", ""], bg="2E4057")
ws5.merge_cells("C9:E9")

steps = [
    ("1", "Verify HMAC-SHA256",     "Using @Value(xtrm.webhook-secret) — throw WebhookAuthenticationException if invalid"),
    ("2", "Parse payload",          "Deserialize raw JSON → XtrmWebhookEvent (externalId, status, providerTransactionId, amount, currency, timestamp)"),
    ("3", "Load transaction",       "Fetch RedemptionTransaction by externalId (= IssuerTransactionId sent in TransferFund)"),
    ("4", "Idempotency check",      "If tx.status already COMPLETED/FAILED/CANCELLED → log and return silently (no duplicate processing)"),
    ("5", "Call WalletService",     "Based on XTRM status: SUCCEEDED→debit(), FAILED/CANCELLED→release()"),
    ("6", "Update transaction",     "Set tx.providerTransactionId = XTRM TransactionId, update tx.status. Save."),
    ("7", "Publish domain event",   "Notify notification service of payout outcome"),
    ("8", "Error handling",         "Wrap entire method in try-catch — ALWAYS return HTTP 200 to XTRM (dead-letter log internal failures)"),
]
for i, (step, action, detail) in enumerate(steps, 10):
    bg = alt_bg(i)
    ws5.cell(row=i, column=1, value=step).fill = hex_fill(BLUE_HEADER)
    ws5.cell(row=i, column=1).font = Font(bold=True, color=WHITE, name="Calibri", size=11)
    ws5.cell(row=i, column=1).alignment = Alignment(horizontal="center", vertical="top")
    ws5.cell(row=i, column=1).border = border
    ws5.cell(row=i, column=2, value=action).fill = hex_fill(bg)
    ws5.cell(row=i, column=2).font = Font(bold=True, name="Calibri", size=10)
    ws5.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws5.cell(row=i, column=2).border = border
    ws5.merge_cells(start_row=i, start_column=3, end_row=i, end_column=5)
    ws5.cell(row=i, column=3, value=detail).fill = hex_fill(bg)
    ws5.cell(row=i, column=3).font = Font(name="Calibri", size=10)
    ws5.cell(row=i, column=3).alignment = Alignment(wrap_text=True, vertical="top")
    ws5.cell(row=i, column=3).border = border
    ws5.row_dimensions[i].height = 30

set_col_widths(ws5, [18, 28, 50, 20, 20])


# ════════════════════════════════════════════════════════════════════════════
# SHEET 6 — DATA GAPS
# ════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("6. Data Gaps")
ws6.sheet_view.showGridLines = False

ws6.merge_cells("A1:E1")
c = ws6["A1"]
c.value = "PART 4 — Data Gaps (Fields required by XTRM not yet in tenXengage model)"
c.fill = hex_fill(DARK_NAVY)
c.font = Font(bold=True, color=WHITE, name="Calibri", size=13)
c.alignment = Alignment(horizontal="center", vertical="center")
ws6.row_dimensions[1].height = 36

header_row(ws6, 2, ["#", "Missing Field", "Required By (XTRM)", "Where to Add", "How to Obtain"], bg=BLUE_HEADER)

gaps = [
    ("G1",  "User.xtrmUserId",                  "CreateUserWallet, TransferFund.RecipientUserId",                       "Add to users table",                               "Returned from CheckUserExist or CreateUser. Cache permanently."),
    ("G2",  "User.xtrmWalletId",                "CreateUserWallet (reference storage)",                                  "Add to users table",                               "Returned from CreateUserWallet. Cache permanently."),
    ("G3",  "User.taxId",                       "CreateUser.TaxId",                                                      "Add encrypted column to user_profiles",            "Collect at onboarding or first CASH redemption KYC. Consider XTRM native tax form flow."),
    ("G4",  "User.dateOfBirth",                 "CreateUser.DateOfBirth",                                                "Add LocalDate to user_profiles",                   "Collect at onboarding."),
    ("G5",  "User.address (full)",              "CreateUser.Address",                                                    "Validate completeness on users table",              "Prompt user to complete at first CASH redemption."),
    ("G6",  "PartnerCompany.xtrmAccountNumber", "TransferFundtoCompany.BeneficiaryAccountNumber",                        "Add to partner_companies table",                   "Obtain from XTRM when company is registered as beneficiary."),
    ("G7",  "PartnerCompany.xtrmWalletId",      "TransferFundtoCompany.BeneficiaryWalletID",                             "Add to partner_companies table",                   "Returned from CreateCompanyWallet."),
    ("G8",  "xtrm.program-id (config)",         "TransferFund.ProgramId",                                                "application.properties / secrets",                 "Create Redemption Store program in XTRM portal. Fetch via GetPrograms."),
    ("G9",  "xtrm.company-wallet-id (config)",  "TransferFund.WalletID",                                                 "application.properties / secrets",                 "Fetch once via GetCompanyWallets after XTRM account is funded."),
    ("G10", "xtrm.payment-method-id (config)",  "TransferFund.PaymentMethodId",                                          "application.properties / secrets",                 "Fetch once via GetUserPaymentMethods. Use XTRM Choice for flexibility."),
    ("G11", "xtrm.webhook-secret (config)",     "XtrmWebhookService HMAC verification",                                  "Secrets manager",                                  "Provided by XTRM when registering webhook endpoint."),
    ("G12", "ISO currency mapping",             "All wallet and payout APIs",                                            "Static map in XtrmAdapter or ClientRedemptionConfig.cashIsoCurrencyCode", "Add cashIsoCurrencyCode (default USD) to ClientRedemptionConfig."),
]

for i, row_data in enumerate(gaps, 3):
    bg = alt_bg(i)
    # # column — orange pill
    c = ws6.cell(row=i, column=1, value=row_data[0])
    c.fill = hex_fill(ORANGE); c.font = Font(bold=True, color=WHITE, name="Calibri", size=10)
    c.alignment = Alignment(horizontal="center", vertical="top"); c.border = border
    for col, val in enumerate(row_data[1:], 2):
        cell = ws6.cell(row=i, column=col, value=val)
        cell.fill = hex_fill(bg)
        cell.font = Font(name="Calibri", size=10)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = border
    ws6.row_dimensions[i].height = 40

set_col_widths(ws6, [6, 38, 45, 42, 62])


# ════════════════════════════════════════════════════════════════════════════
# SHEET 7 — OPEN QUESTIONS
# ════════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("7. Open Questions")
ws7.sheet_view.showGridLines = False

ws7.merge_cells("A1:F1")
c = ws7["A1"]
c.value = "PART 5 — Open Questions & Ambiguities"
c.fill = hex_fill(DARK_NAVY)
c.font = Font(bold=True, color=WHITE, name="Calibri", size=13)
c.alignment = Alignment(horizontal="center", vertical="center")
ws7.row_dimensions[1].height = 36

header_row(ws7, 2, ["#", "Title", "Priority", "Description", "Required Action", "Owner"], bg=BLUE_HEADER)

oqs = [
    ("OQ-1",  "Wrong Endpoint in Document",      "HIGH",   "The document specifies POST /v1/rewards as the XTRM payout endpoint. The actual XTRM XAPI uses POST /API/V4/Fund/TransferFund with a completely different request schema. Code generated from the document will NOT work against real XTRM.", "Confirm with XTRM account team before XtrmAdapter is coded.", ""),
    ("OQ-2",  "apidoc.xtrm.com Is Inaccessible", "HIGH",   "The URL provided (apidoc.xtrm.com) renders as a JavaScript SPA that cannot be crawled. Referenced section may contain different schemas than the legacy XAPI console used in this analysis.", "Access apidoc.xtrm.com with an authenticated browser session and compare endpoint schemas.", ""),
    ("OQ-3",  "ProgramId Purpose and Selection",  "MEDIUM", "XTRM Programs categorise transactions (SPIFF, MDF, rebate, etc). ProgramId is required for TransferFund. Unclear which program type maps to a redemption payout.", "Create a dedicated Redemption Store program in the XTRM portal. Confirm TransactionCategoryID with XTRM.", ""),
    ("OQ-4",  "PaymentMethodId and Xtrm Choice",  "MEDIUM", "XTRM supports multiple payout methods (bank transfer, prepaid Visa, PayPal). Should tenXengage use Xtrm Choice (recipient picks) or direct-to-bank?", "Confirm payout method strategy with product owner. Xtrm Choice is lowest friction — populate UserPrepaidVisaEmailID with recipient email.", ""),
    ("OQ-5",  "XTRM Webhook Payload Schema",      "HIGH",   "XTRM's exact webhook event field names are not publicly documented. The document assumes externalId, status, providerTransactionId. Actual field names need verification.", "Register webhook in XTRM sandbox, trigger a test payout, and capture the real payload before coding XtrmWebhookService.", ""),
    ("OQ-6",  "Company Payout Flow",              "MEDIUM", "TransferFundtoCompany requires BeneficiaryAccountNumber (the partner company's XTRM SPN). Unclear if partner companies are registered on XTRM as issuers or just beneficiary users.", "Clarify with XTRM. If company accounts not supported, route company payouts through the representative user's individual account.", ""),
    ("OQ-7",  "TaxId Collection and Compliance",  "HIGH",   "TaxId (SSN/EIN) is sensitive PII subject to regulatory controls. Collecting it in tenXengage introduces 1099-K compliance burden. XTRM can alternatively collect tax forms natively.", "Evaluate XTRM's native tax form flow. If usable, make CreateUser.TaxId optional and avoid storing SSN in tenXengage.", ""),
    ("OQ-8",  "Idempotency on TransferFund",      "HIGH",   "If XtrmAdapter.submitPayout() is retried (VendorUnavailableException), will XTRM process the same IssuerTransactionId a second time (double-payment)?", "Confirm XTRM idempotency guarantee on IssuerTransactionId. If not guaranteed, add pre-retry guard checking whether tx is already PROCESSING.", ""),
    ("OQ-9",  "XTRM Wallet Balance Monitoring",   "MEDIUM", "The XTRM source wallet must have sufficient funds. No mechanism to monitor or auto-replenish it. If wallet runs dry, all redemptions fail.", "Add a scheduled job calling GetCompanyWallets to check balance. Alert when below a threshold.", ""),
    ("OQ-10", "BATCH Processing Mode",            "MEDIUM", "BATCH mode sets tx.status=PENDING and schedules payout for next batch run. The batch runner is not defined anywhere in the document.", "Define a @Scheduled batch runner in RedemptionOrchestrationService querying PENDING transactions and calling routeToVendor().", ""),
    ("OQ-11", "Duplicate Email Notifications",    "LOW",    "When EmailNotification=true in TransferFund, XTRM sends its own email. The tenXengage notification service may send a second one via the webhook domain event.", "Set EmailNotification=false and rely on tenXengage notifications, OR inform the user they will receive two emails.", ""),
    ("OQ-12", "Multi-Currency Support",           "MEDIUM", "CurrencyType.CASH is currently mapped to USD. For partners in other regions (India, UK etc.), this mapping breaks.", "Add cashIsoCurrencyCode (String, ISO 4217, default USD) to ClientRedemptionConfig.", ""),
]

for i, (num, title, priority, desc, action, owner) in enumerate(oqs, 3):
    bg = alt_bg(i)
    c = ws7.cell(row=i, column=1, value=num)
    c.fill = hex_fill(BLUE_HEADER); c.font = Font(bold=True, color=WHITE, name="Calibri", size=10)
    c.alignment = Alignment(horizontal="center", vertical="top"); c.border = border
    ws7.cell(row=i, column=2, value=title).fill = hex_fill(bg)
    ws7.cell(row=i, column=2).font = Font(bold=True, name="Calibri", size=10)
    ws7.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws7.cell(row=i, column=2).border = border
    priority_cell(ws7, i, 3, priority)
    for col, val in enumerate([desc, action, owner], 4):
        cell = ws7.cell(row=i, column=col, value=val)
        cell.fill = hex_fill(bg)
        cell.font = Font(name="Calibri", size=10)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = border
    ws7.row_dimensions[i].height = 55

set_col_widths(ws7, [7, 32, 12, 55, 52, 20])


# ════════════════════════════════════════════════════════════════════════════
# SHEET 8 — REST CONTROLLERS
# ════════════════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("8. REST Controllers")
ws8.sheet_view.showGridLines = False

ws8.merge_cells("A1:E1")
c = ws8["A1"]
c.value = "PART 6 — REST Controllers (XTRM-Relevant Endpoints)"
c.fill = hex_fill(DARK_NAVY)
c.font = Font(bold=True, color=WHITE, name="Calibri", size=13)
c.alignment = Alignment(horizontal="center", vertical="center")
ws8.row_dimensions[1].height = 36

header_row(ws8, 2, ["Controller", "Endpoint", "HTTP", "XTRM Involvement", "Direction"], bg=BLUE_HEADER)

controllers = [
    ("RedemptionController",        "/api/v1/redemption/requests",           "POST", "OrchestrationService → XtrmAdapter.submitPayout() if CASH + INSTANT",           "Outbound → XTRM"),
    ("RedemptionController",        "/api/v1/redemption/requests/company",   "POST", "Same flow using TransferFundtoCompany",                                          "Outbound → XTRM"),
    ("RedemptionController",        "/api/v1/redemption/requests/{id}/approve","POST","OrchestrationService.approveRedemption() → XtrmAdapter.submitPayout()",         "Outbound → XTRM"),
    ("RedemptionWebhookController", "/api/v1/redemption/webhook/xtrm",       "POST", "Inbound from XTRM, handled by XtrmWebhookService (HMAC-SHA256 secured)",         "Inbound ← XTRM"),
    ("WalletController",            "/api/v1/wallets/me",                    "GET",  "No XTRM call — returns internal wallet state only",                              "Internal only"),
    ("WalletController",            "/api/v1/wallets/company/{companyId}",   "GET",  "No XTRM call — returns internal wallet state only",                              "Internal only"),
]

dir_colors = {"Outbound → XTRM": ("1A5276", WHITE), "Inbound ← XTRM": ("117A65", WHITE), "Internal only": ("626567", WHITE)}
for i, (ctrl, ep, method, inv, direction) in enumerate(controllers, 3):
    bg = alt_bg(i)
    for col, val in enumerate([ctrl, ep, method, inv], 1):
        cell = ws8.cell(row=i, column=col, value=val)
        cell.fill = hex_fill(bg)
        cell.font = Font(name="Calibri", size=10)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = border
    ws8.cell(row=i, column=1).font = Font(name="Courier New", size=9, color="1A5276")
    ws8.cell(row=i, column=2).font = Font(name="Courier New", size=9, color="117A65")
    d_bg, d_fg = dir_colors.get(direction, (WHITE, "000000"))
    c = ws8.cell(row=i, column=5, value=direction)
    c.fill = hex_fill(d_bg); c.font = Font(bold=True, color=d_fg, name="Calibri", size=10)
    c.alignment = Alignment(horizontal="center", vertical="top"); c.border = border
    ws8.row_dimensions[i].height = 35

set_col_widths(ws8, [30, 45, 8, 60, 20])

# ── Save ────────────────────────────────────────────────────────────────────
out_path = r"C:\Users\TenXengage\Development\TenXEngage-New\Redemption Store Functionality\TenXEngage_XTRM_Mapping_Details.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
